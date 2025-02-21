from flask import Flask, render_template, request, redirect, session, make_response, send_file
import sqlite3
import datetime
import re
import csv
import io
from openpyxl import Workbook
from fpdf import FPDF

app = Flask(__name__)
app.secret_key = "super_secret_key"

# Global variable to store blocked IPs
blocked_ips = set()

# Login Page
@app.route("/", methods=["GET", "POST"])
def login():
    ip = request.remote_addr
    # Check if the IP is already blocked before processing login
    if ip in blocked_ips:
        print(f"DEBUG: Blocked attempt from {ip}")
        return "Your IP has been blocked due to suspicious activity.", 403

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        timestamp = datetime.datetime.now()

        # Log the attempt (regardless of outcome)
        # We'll log "Success" or "Failed" based on the result.
        conn = sqlite3.connect("users.db")
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
        user = c.fetchone()
        conn.close()

        if user:
            log_attempt(ip, username, timestamp, "Success")
            return f"Welcome {username}!"
        else:
            log_attempt(ip, username, timestamp, "Failed")
            return "Invalid credentials."

    return render_template("login.html")


# Function to log login attempts and trigger brute-force check
def log_attempt(ip, username, timestamp, status):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("INSERT INTO login_attempts (ip, username, timestamp, status) VALUES (?, ?, ?, ?)",
              (ip, username, timestamp, status))
    conn.commit()
    conn.close()
    check_brute_force(ip)


# Function to check brute-force attack
def check_brute_force(ip):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    # Count failed attempts in the last 5 minutes for the given IP
    c.execute("SELECT COUNT(*) FROM login_attempts WHERE ip=? AND timestamp >= datetime('now', '-5 minutes') AND status='Failed'", (ip,))
    attempts = c.fetchone()[0]
    conn.close()
    
    print(f"DEBUG: {ip} has {attempts} failed attempts in the last 5 minutes.")
    
    if attempts > 5:
        print(f"DEBUG: Brute force detected from {ip}. Blocking...")
        block_ip(ip)


# Function to block IP
def block_ip(ip):
    blocked_ips.add(ip)
    print(f"DEBUG: IP {ip} has been added to the blocked list.")


# Utility function to fetch login attempts (for downloads)
def get_login_attempts():
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("SELECT * FROM login_attempts")
    attempts = c.fetchall()
    conn.close()
    return attempts


# CSV Download Route
@app.route("/download/csv")
def download_csv():
    attempts = get_login_attempts()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "IP", "Username", "Timestamp", "Status"])
    for attempt in attempts:
        writer.writerow(attempt)
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=login_attempts.csv"
    response.headers["Content-type"] = "text/csv"
    return response


# Excel Download Route
@app.route("/download/excel")
def download_excel():
    attempts = get_login_attempts()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "IP", "Username", "Timestamp", "Status"])
    for attempt in attempts:
        sheet.append(attempt)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="login_attempts.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# PDF Download Route
@app.route("/download/pdf")
def download_pdf():
    attempts = get_login_attempts()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Login Attempts", ln=True, align="C")
    pdf.ln(10)
    pdf.cell(20, 10, "ID", 1)
    pdf.cell(40, 10, "IP", 1)
    pdf.cell(40, 10, "Username", 1)
    pdf.cell(50, 10, "Timestamp", 1)
    pdf.cell(40, 10, "Status", 1)
    pdf.ln()
    for attempt in attempts:
        pdf.cell(20, 10, str(attempt[0]), 1)
        pdf.cell(40, 10, str(attempt[1]), 1)
        pdf.cell(40, 10, str(attempt[2]), 1)
        pdf.cell(50, 10, str(attempt[3]), 1)
        pdf.cell(40, 10, str(attempt[4]), 1)
        pdf.ln()
    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="login_attempts.pdf", mimetype="application/pdf")


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5001, debug=True)
